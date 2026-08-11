"""
generate_index_coverage_report.py - Turns Search Console's Page Indexing report
into the client-ready "Index Coverage Report" workbook the team builds by hand
(reference: Index-coverage-report-<domain>.xlsx).

The official Search Console API does not expose the Page Indexing report's
per-reason URL lists at all - this is a real, documented gap, not something
this script works around by choice. The URL lists have to come from somewhere
that already scraped them: gsc_audit.py's capture_index_coverage_urls()
(Selenium, driving the same already-authenticated GSC browser session every
other GSC tool in this app already uses) is the intended source, but this
module only cares about the shape: {reason_name: [url, ...], ...}. Anything
that can produce that dict can feed this.

For every URL, this:
  - Checks its REAL current HTTP status + follows redirects (Search Console's
    own reason was recorded whenever it last crawled - can be stale; a page
    listed under "Crawled - currently not indexed" may actually 301 or 404
    TODAY, which changes what the real fix is).
  - Classifies Indexability / Indexability Status from that live check (not
    invented - "Redirected" only when the live check found a redirect,
    "Client Error" only from a real 4xx, etc.).
  - For "Not found (404)" specifically, suggests a real redirect target - the
    site's own sitemap's best-matching live URL by slug similarity, never a
    fabricated URL. No match found -> left blank, not guessed.

Output: <domain>'s Index Coverage Report.xlsx - "Index" summary tab (reasons
that need real action, with a one-line fix each) + one tab per reason
actually present (title + definition row + data), matching the reference
workbook's layout exactly.

Run:
    python generate_index_coverage_report.py --domain example.com \
        --reason-urls reason_urls.json --out "Index Coverage Report.xlsx"
"""
import os
import re
import sys
import json
import time
import argparse
import urllib.parse
import urllib.request
import urllib.error
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

ROOT = Path(__file__).parent.resolve()
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import generate_seo_onpage_phase2 as onpage2   # live page fetch (title/canonical/meta robots)
import generate_geo_report as georpt           # get_sitemap_urls - real sitemap discovery


def log(msg):
    try:
        print(msg, flush=True)
    except UnicodeEncodeError:
        enc = sys.stdout.encoding or "ascii"
        print(str(msg).encode(enc, errors="replace").decode(enc), flush=True)


# --------------------------------------------------------------------------- #
# Reason definitions - Google's own documented Page Indexing reasons (Search
# Console Help: "Why pages aren't indexed"), not just the subset that happened
# to show up in any one reference report. A reason not in this table (Google
# adds new ones occasionally) still gets its own tab with a generic fallback
# definition, same as the reference workbook does for its one non-standard
# reason - never silently dropped.
# --------------------------------------------------------------------------- #
# needs_action: whether this reason belongs in the "Index" summary tab at all
# (healthy/expected-by-design reasons like "Page with redirect" or "Alternate
# page with proper canonical tag" still get their own detail tab, just not a
# row in the action-needed summary - matches the reference workbook, whose
# Index tab lists only 4 of its 9 reasons).
REASON_INFO = {
    "Discovered - currently not indexed": {
        "definition": "Google has found this URL (via a link or sitemap) but has not crawled it yet, "
                      "so it cannot appear in search. Often a sign of crawl-budget limits, weak internal "
                      "linking, or thin value.",
        "needs_action": True,
        "fix_summary": "Strengthen internal linking to these pages and confirm they carry real value - "
                       "weak/thin pages rarely get crawled.",
        "default_action": "Improve internal linking and page value so Google prioritizes crawling this URL.",
    },
    "Crawled - currently not indexed": {
        "definition": "Google crawled the page but decided not to index it, so it will not appear in "
                      "search. Usually means the content is thin, near-duplicate, or low value in "
                      "Google's judgement.",
        "needs_action": True,
        "fix_summary": "Google crawled the page but chose not to index it. Enhance the content to make "
                       "it more useful, original, and distinct from other pages.",
        "default_action": "Need to index this page",
    },
    "Soft 404": {
        "definition": "The server returns 200 OK, telling Google the page works, but the content is "
                      "empty, thin, or reads like 'not found', so Google sees no real value. Common "
                      "causes: blank or thin pages (empty tag pages, out-of-stock products with no "
                      "alternatives), deleted URLs redirected to the homepage instead of returning 404, "
                      "or content hidden from bots by JavaScript or CSS. Fix it by enriching the content "
                      "if the page should live, or removing it properly with a 404 or 410, or a 301 to "
                      "a relevant page. NOTE on this tab's own Status Code/Indexability columns: they "
                      "reflect the raw server response from a quick check, same as every other tab - a "
                      "'200 / Indexable' row here does NOT mean the page is actually fine. Google's own "
                      "crawler renders JavaScript and judges the real visible content (which is exactly "
                      "why it flagged this URL as Soft 404 in the first place, even though the server "
                      "says 200) - this report's fast check can't replicate that render step, so treat a "
                      "200 here as 'the server responds' only, not as 'this page is actually healthy'.",
        "needs_action": True,
        "fix_summary": "Enrich the content if the page should exist, or remove it properly (404/410, "
                       "or a 301 to a relevant page).",
        "default_action": "Enrich the content, or remove it properly (404/410, or 301 to a relevant page)",
    },
    "Not found (404)": {
        "definition": "The URL returns 404 Not Found, so it cannot be indexed or ranked. Any inbound "
                      "links and past ranking value are lost unless the URL is redirected to a relevant "
                      "live page.",
        "needs_action": True,
        "fix_summary": "The URL returns 404. Redirect it to the closest relevant live page, or restore "
                       "the page if it should exist.",
        "default_action": None,   # per-row: filled from the real redirect suggestion, see build_report()
    },
    "Not found (410)": {
        "definition": "The URL deliberately returns 410 Gone, telling Google the page was intentionally "
                      "and permanently removed. Stronger than a 404 - only use this when the page should "
                      "never come back.",
        "needs_action": False,
        "fix_summary": "Working as intended if this removal was deliberate; redirect it instead if the "
                       "content moved rather than disappeared.",
        "default_action": "No action needed if deliberate - otherwise redirect to a relevant live page",
    },
    "Duplicate without user-selected canonical": {
        "definition": "Google found multiple near-identical URLs and no canonical tag told it which one "
                      "you prefer, so it picked one itself (which may not be the one you want ranked).",
        "needs_action": True,
        "fix_summary": "Add a self-referencing canonical tag on the version you want indexed.",
        "default_action": "Add a canonical tag pointing to the preferred version of this page",
    },
    "Duplicate, Google chose different canonical than user": {
        "definition": "You set a canonical but Google overrode it and chose a different page. Usually "
                      "Google trusts another version more; if your choice should win, strengthen its "
                      "internal links, sitemap entry, and canonical.",
        "needs_action": False,
        "fix_summary": "No action needed unless the page Google chose isn't the one you actually want "
                       "ranked - if so, strengthen the preferred version's internal links and sitemap entry.",
        "default_action": "No Action Needed",
    },
    "Duplicate, submitted URL not selected as canonical": {
        "definition": "You submitted this exact URL (e.g. via sitemap), but Google indexed a different "
                      "duplicate/variant of it as the canonical instead.",
        "needs_action": False,
        "fix_summary": "No action needed unless the canonical Google picked isn't the version you want "
                       "ranked - if so, strengthen this URL's own signals instead.",
        "default_action": "No Action Needed",
    },
    "Alternate page with proper canonical tag": {
        "definition": "This is a duplicate or variant (such as a paginated or parameter URL) that "
                      "correctly points to its canonical version, so Google indexes the canonical "
                      "instead. Working as intended.",
        "needs_action": False,
        "fix_summary": "Working as intended - no action needed.",
        "default_action": "No action needed",
    },
    "Blocked by robots.txt": {
        "definition": "robots.txt tells Googlebot not to crawl this URL, so Google cannot read it (it "
                      "may still show as a bare link). Fine when intentional; a problem only if it "
                      "blocks pages you want ranked.",
        "needs_action": False,
        "fix_summary": "No action needed for pages you don't want crawled - otherwise remove the "
                       "robots.txt rule blocking this URL.",
        "default_action": "Not Important Page - No action needed",
    },
    "Blocked due to unauthorized request (401)": {
        "definition": "The page returned 401 Unauthorized to Googlebot, so it couldn't be crawled or "
                      "indexed. Usually login-gated content, or a WAF/bot-check wrongly challenging "
                      "Googlebot's own requests.",
        "needs_action": True,
        "fix_summary": "Remove the login/auth requirement for pages you want indexed, or confirm this "
                       "page should genuinely stay gated.",
        "default_action": "Check whether Googlebot should have access to this page",
    },
    "Blocked due to access forbidden (403)": {
        "definition": "The page returned 403 Forbidden to Googlebot, so it couldn't be crawled or "
                      "indexed. Often a firewall/bot-protection rule blocking crawlers, not just users.",
        "needs_action": True,
        "fix_summary": "Whitelist Googlebot in the firewall/bot-protection rule blocking this page, if "
                       "the page should be indexed.",
        "default_action": "Check server/firewall rules blocking Googlebot from this page",
    },
    "Blocked due to other 4xx issue": {
        "definition": "The page returned some other 4xx client-error status (not 401/403/404) to "
                      "Googlebot, so it couldn't be indexed.",
        "needs_action": True,
        "fix_summary": "Inspect the URL directly to see the exact status code and fix the server-side "
                       "issue causing it.",
        "default_action": "Inspect the URL in Search Console for the exact error and fix it",
    },
    "Blocked by page removal tool": {
        "definition": "Someone used Search Console's Removals tool to temporarily hide this URL from "
                      "search results. Temporary by design - it doesn't remove the page from the index "
                      "permanently.",
        "needs_action": False,
        "fix_summary": "No action needed if this removal was intentional - it expires automatically "
                       "after ~6 months unless renewed.",
        "default_action": "No action needed if intentional",
    },
    "Page with redirect": {
        "definition": "This URL redirects to another URL, so the redirecting URL is not indexed; the "
                      "destination is what can rank. Normal for moved or canonical URLs. Only redirect "
                      "chains (several hops) or loops hurt performance and need fixing.",
        "needs_action": False,
        "fix_summary": "Working as intended for a normal single-hop redirect - only fix if this is part "
                       "of a redirect chain or loop.",
        "default_action": "No Action Needed",
    },
    "Redirect error": {
        "definition": "Googlebot could not follow this URL's redirect - a redirect loop, chain too long, "
                      "a redirect to an empty URL, or the redirect URL exceeded the max length.",
        "needs_action": True,
        "fix_summary": "Fix the redirect - remove loops/long chains and redirect straight to the final "
                       "live destination in one hop.",
        "default_action": "Fix the redirect to point directly at the final destination in one hop",
    },
    "Server error (5xx)": {
        "definition": "The server returned a 5xx error when Googlebot requested this URL, so it "
                      "couldn't be crawled or indexed. Usually a server-side problem, not the page "
                      "itself.",
        "needs_action": True,
        "fix_summary": "Investigate the server error (check server logs / hosting) - this is a hosting "
                       "issue, not a content issue.",
        "default_action": "Investigate the server error with your hosting provider",
    },
    "Excluded by 'noindex' tag": {
        "definition": "A meta robots or X-Robots-Tag 'noindex' directive on this page explicitly tells "
                      "Google not to index it. Working as intended if deliberate; a real problem if this "
                      "page should actually rank.",
        "needs_action": True,
        "fix_summary": "Remove the noindex tag if this page should be indexed and ranked; leave it if "
                       "the exclusion is intentional.",
        "default_action": "Need to remove noindex tag from important pages",
    },
    "Not indexed due to legal removal": {
        "definition": "This URL was removed from Google's index due to a legal request (e.g. a valid "
                      "DMCA takedown), not a technical or content issue.",
        "needs_action": False,
        "fix_summary": "No action available from the site side - this is a legal removal on Google's end.",
        "default_action": "No action available - legal removal",
    },
}

# Fallback for any reason Google shows that isn't in the table above (new
# reason types, or a name variant) - same generic text the reference workbook
# itself uses, never silently skipped.
_GENERIC_REASON_INFO = {
    "definition": "This reason is not in the standard list. Inspect the URL in Search Console to see "
                  "Google's own explanation, then act accordingly.",
    "needs_action": True,
    "fix_summary": "Inspect a sample URL in Search Console's URL Inspection tool for the specific reason.",
    "default_action": "Inspect in Search Console and act accordingly",
}


_UNICODE_PUNCT_NORMALIZE = {
    "‘": "'", "’": "'", "“": '"', "”": '"',   # curly quotes
    "–": "-", "—": "-",                                 # en/em dash
}
_REASON_LOOKUP = {}
for _name, _info in REASON_INFO.items():
    _REASON_LOOKUP[_name.translate(str.maketrans(_UNICODE_PUNCT_NORMALIZE))] = _info


def _reason_info(reason):
    """Looks up REASON_INFO tolerantly - confirmed live, GSC's real exported
    reason names use curly quotes/en-dashes ("Excluded by 'noindex' tag",
    "Discovered - currently not indexed" with typographic punctuation) while
    this table was typed with plain ASCII ones; a straight equality lookup
    would silently miss every one of those and fall back to the generic
    entry despite having a real, tailored definition for it."""
    if reason in REASON_INFO:
        return REASON_INFO[reason]
    normalized = reason.translate(str.maketrans(_UNICODE_PUNCT_NORMALIZE))
    return _REASON_LOOKUP.get(normalized, _GENERIC_REASON_INFO)


# --------------------------------------------------------------------------- #
# Live per-URL status/redirect check - never trusts GSC's own (possibly
# stale) recorded reason on its own; every URL gets checked fresh.
# --------------------------------------------------------------------------- #
_UA = "Mozilla/5.0 IndexCoverageReportBot"


def _safe_url(url):
    try:
        parts = urllib.parse.urlsplit(url)
        path = urllib.parse.quote(parts.path, safe="/%")
        query = urllib.parse.quote(parts.query, safe="=&%")
        return urllib.parse.urlunsplit((parts.scheme, parts.netloc, path, query, parts.fragment))
    except Exception:
        return url


class _NoAutoRedirect(urllib.request.HTTPRedirectHandler):
    """Disables urllib's normal silent-follow-and-hide-the-chain behavior -
    returning None from redirect_request makes urllib raise the 3xx as an
    HTTPError (with .code and a Location header) instead of quietly moving
    on, so each hop can be recorded and inspected individually."""
    def redirect_request(self, *a, **kw):
        return None


def follow_redirect_chain(url, max_hops=10, timeout=15):
    """Walks the FULL redirect chain hop by hop - never trusts the "did it
    eventually land on a 200" summary alone, since a page GSC lists as "Page
    with redirect" or "Not found (404)" needs the real story: is it a clean
    single 301 straight to a live page, a chain of several hops, a loop, or
    a redirect that itself dead-ends in a 404? That distinction is exactly
    what determines the real fix (leave it, consolidate the chain, or
    redirect it directly to a genuinely live page).

    Returns {"chain": [{"url":, "status":}, ...] (every hop actually
    visited, in order), "final_url":, "final_status":, "hop_count":
    (0 = no redirect at all), "html": <final page's HTML, if any>,
    "loop": bool}. Never raises - a network failure on any hop ends the
    chain there with status=None for that hop rather than crashing the run."""
    opener = urllib.request.build_opener(_NoAutoRedirect)
    chain = []
    seen = set()
    current = url
    html = ""
    loop = False
    for _ in range(max_hops):
        if current in seen:
            loop = True
            break
        seen.add(current)
        # 429 (Too Many Requests) gets a couple of short, backed-off retries
        # before being recorded as a real status - confirmed live this can be
        # SELF-inflicted: this report checks many URLs on the same origin
        # concurrently (see process_reason()'s thread pool), which can trip a
        # site's own rate limiter even though the page itself is perfectly
        # healthy. Honors a real Retry-After header when the server sends one.
        rate_limited = False
        for retry in range(3):
            try:
                req = urllib.request.Request(_safe_url(current), headers={"User-Agent": _UA}, method="GET")
                with opener.open(req, timeout=timeout) as r:
                    chain.append({"url": current, "status": r.status})
                    html = r.read().decode("utf-8", "ignore")
                rate_limited = False
                break   # 2xx with no redirect handler firing - chain ends here
            except urllib.error.HTTPError as e:
                if e.code == 429 and retry < 2:
                    rate_limited = True
                    try:
                        wait = float(e.headers.get("Retry-After", 2 * (retry + 1)))
                    except (TypeError, ValueError):
                        wait = 2 * (retry + 1)
                    time.sleep(min(wait, 5))
                    continue
                rate_limited = False
                chain.append({"url": current, "status": e.code})
                if e.code in (301, 302, 303, 307, 308):
                    loc = e.headers.get("Location")
                    if not loc:
                        break
                    current = urllib.parse.urljoin(current, loc)
                    rate_limited = "redirect"
                    break
                try:
                    html = e.read().decode("utf-8", "ignore")
                except Exception:
                    html = ""
                break
            except Exception as e:
                rate_limited = False
                log(f"   [warn] could not check {current}: {type(e).__name__}: {e}")
                chain.append({"url": current, "status": None})
                break
        if rate_limited == "redirect":
            continue
        break
    final = chain[-1] if chain else {"url": url, "status": None}
    return {"chain": chain, "final_url": final["url"], "final_status": final["status"],
            "hop_count": len(chain) - 1, "html": html, "loop": loop}


def check_live_status(url):
    """Real current HTTP status + final URL after following redirects -
    thin wrapper over follow_redirect_chain() for callers that only need
    the immediate/final picture, not the full hop-by-hop chain."""
    result = follow_redirect_chain(url, max_hops=10)
    return {"status": result["final_status"], "final_url": result["final_url"], "html": result["html"]}


def classify_indexability(url, live):
    """(Indexability, Indexability Status, Canonical Target) - derived from
    the REAL live check just performed, never from GSC's own possibly-stale
    reason. Canonical Target is only populated when Indexability Status is
    "Canonicalised" (the live page itself points elsewhere via its own
    canonical tag) - None otherwise, so callers can show "what does this
    page's canonical actually say right now" instead of just flagging that
    one exists."""
    status = live.get("status")
    final_url = live.get("final_url") or url
    if status is None:
        return "Unknown", "Could not check - please verify manually", None
    if 300 <= status < 400 or (final_url and final_url.rstrip("/") != url.rstrip("/")):
        return "Non-Indexable", "Redirected", None
    if status in (401, 403):
        return "Non-Indexable", "Blocked", None
    if status == 404 or status == 410:
        return "Non-Indexable", "Client Error", None
    if status >= 500:
        return "Non-Indexable", "Server Error", None
    if status == 200:
        html = live.get("html") or ""
        pd = onpage2._parse_html(html, final_url, status) if html else None
        if pd:
            robots_meta = re.search(r'<meta[^>]+name=["\']robots["\'][^>]+content=["\']([^"\']*noindex[^"\']*)',
                                    html, re.I)
            if robots_meta:
                return "Non-Indexable", "noindex", None
            canon = pd.get("canonical") or ""
            m = re.search(r'href="([^"]+)"', canon)
            if m and m.group(1).rstrip("/") != url.rstrip("/"):
                return "Non-Indexable", "Canonicalised", m.group(1)
        return "Indexable", "-", None
    return "Unknown", "Could not check - please verify manually", None


_slug_word_re = re.compile(r"[a-z0-9]+")


def _slug_words(url):
    path = urllib.parse.urlparse(url).path
    return set(_slug_word_re.findall(path.lower()))


# Static resource files GSC's Page Indexing report can list right alongside
# real content pages (a theme's own JS/CSS, images, fonts, sitemap/robots
# files) - confirmed live via a real report recommending "redirect the
# broken page to → homepage" for a WordPress emoji script
# (wp-emoji-release.min.js) returning 403. That's nonsensical: a JS file
# isn't a content page competing for a search ranking, it doesn't need to be
# indexed, and redirecting it to the homepage doesn't fix anything - the
# actual question (why does a static asset 403, if that's even a real
# problem) is a totally different, usually non-SEO issue.
_STATIC_ASSET_EXTS = (
    ".js", ".mjs", ".css", ".map", ".json", ".xml",
    ".woff", ".woff2", ".ttf", ".eot", ".otf",
    ".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".ico", ".bmp", ".avif",
)


def _is_static_asset(url):
    path = urllib.parse.urlparse(url).path.lower()
    return path.endswith(_STATIC_ASSET_EXTS)


def _ranked_redirect_candidates(dead_url, sitemap_urls, homepage_url):
    """Sitemap URLs ranked by slug-word overlap with the dead URL (best
    first), homepage appended last as the universal fallback - never a
    fabricated URL, only real pages the site's own sitemap lists."""
    dead_words = _slug_words(dead_url)
    scored = []
    if dead_words:
        for cand in sitemap_urls:
            cand_words = _slug_words(cand)
            if not cand_words:
                continue
            score = len(dead_words & cand_words) / max(len(dead_words), 1)
            if score >= 0.34:
                scored.append((score, cand))
    scored.sort(key=lambda t: -t[0])
    ranked = [c for _, c in scored]
    if homepage_url not in ranked:
        ranked.append(homepage_url)
    return ranked


def find_live_redirect_target(dead_url, sitemap_urls, homepage_url, max_candidates=4):
    """The best REAL, CONFIRMED-LIVE page to redirect a broken/dead-ending
    URL to - not just the best slug match, but the best slug match that
    itself actually resolves cleanly (live-checked, same as every other URL
    in this report - a suggested redirect target that turns out to be
    ANOTHER dead link or another redirect chain would just move the problem,
    not fix it). Tries candidates best-match-first, stops at the first one
    that resolves to a clean 200; the homepage is always the last resort and
    checked too (a real hosting outage would otherwise silently recommend a
    homepage that's ALSO down, without saying so).

    Returns the winning follow_redirect_chain() result dict, or None if not
    even the homepage could be confirmed live (report this honestly rather
    than recommending an unconfirmed URL)."""
    candidates = _ranked_redirect_candidates(dead_url, sitemap_urls, homepage_url)[:max_candidates]
    for cand in candidates:
        result = follow_redirect_chain(cand, max_hops=5)
        if result["final_status"] == 200:
            return result
    return None


# --------------------------------------------------------------------------- #
# Output workbook - matches the reference layout: title row, definition row,
# header row, then data - per reason tab, plus an "Index" summary tab.
# --------------------------------------------------------------------------- #
TITLE_FONT = Font(bold=True, size=13)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
NO_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _safe_sheet_name(name, used):
    safe = re.sub(r"[\[\]:\\/?*]", "_", name)[:31] or "Sheet"
    base, n = safe, 1
    while safe in used:
        n += 1
        safe = f"{base[:28]}_{n}"
    used.add(safe)
    return safe


def build_report(domain, reason_rows, out_path, brand=None):
    """reason_rows: {reason: [row, ...], ...} - already live-checked (see
    process_reason() below, which is how main() builds this). Each row is
    either a "simple" row ({"url":, "last_crawled":, "status":,
    "indexability":, "indexability_status":, "canonical_target":, "action":})
    or, for "Not found (404)"/"Page with redirect", a "redirect" row
    ({"url":, "last_crawled":, "current_status":, "redirects_to":,
    "destination_status":, "hop_count":, "recommendation":,
    "redirect_target":, "target_status":})."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names = set()

    # --- Index summary tab (action-needed reasons only, matching reference) ---
    ws = wb.create_sheet(_safe_sheet_name("Index", used_names))
    ws.append(["Index Coverage Report"])
    ws["A1"].font = TITLE_FONT
    ws.append([domain])
    ws.append(["Every page-indexing reason from Search Console except healthy indexed pages. "
               "Each tab opens with a definition of the reason."])
    ws.append(["Reason", "Action", "Recommended fix"])
    for c in ws[4]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for reason in reason_rows:
        info = _reason_info(reason)
        if info["needs_action"]:
            ws.append([reason, "Action needed", info["fix_summary"]])
    for col, width in zip("ABC", (45, 16, 90)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = WRAP

    # --- One tab per reason actually found ---
    for reason, rows in reason_rows.items():
        info = _reason_info(reason)
        sheet_name = _safe_sheet_name(reason, used_names)
        ws = wb.create_sheet(sheet_name)
        ws.append([reason])
        ws["A1"].font = TITLE_FONT
        ws.append([info["definition"]])
        ws["A2"].alignment = WRAP
        # Scales with the definition's own length instead of a fixed 60 -
        # confirmed the Soft 404 entry's definition (which carries an extra
        # caveat paragraph about this report's own Status Code/Indexability
        # columns) runs noticeably longer than the others and would
        # otherwise get visually cut off in Excel's fixed-height wrap.
        # ~95 chars/line at this column width, ~15px/line - never smaller
        # than the original 60 for the shorter definitions.
        ws.row_dimensions[2].height = max(60, 15 * (len(info["definition"]) // 95 + 2))

        is_redirect_sheet = bool(rows) and rows[0].get("kind") == "redirect"
        if is_redirect_sheet:
            # Full chain analysis - not just "is it a 200 now", but the whole
            # story: this URL's own current status, exactly where it
            # currently redirects to, what THAT resolves to, how many hops
            # that took, then the suggestion as 3 separate columns
            # (recommendation / target URL / target's own live status)
            # rather than one text blob, so the target URL is directly
            # usable instead of buried in a sentence.
            headers = ["Address", "Last Crawled", "Current Status (Live)", "Redirects To",
                      "Destination Status", "Hops", "Recommendation",
                      "Suggested Redirect To", "Suggested Target Status"]
            ws.append(headers)
            for c in ws[3]:
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
            for r in rows:
                ws.append([r["url"], r.get("last_crawled") or "-", r.get("current_status") or "-",
                          r.get("redirects_to") or "-", r.get("destination_status") or "-",
                          r.get("hop_count", ""), r.get("recommendation", ""),
                          r.get("redirect_target") or "-", r.get("target_status") or "-"])
            widths = (55, 14, 18, 55, 18, 8, 45, 55, 22)
        else:
            headers = ["Address", "Last Crawled", "Status Code", "Indexability",
                      "Indexability Status", "Canonical Points To", "Action"]
            ws.append(headers)
            for c in ws[3]:
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
            for r in rows:
                ws.append([r["url"], r.get("last_crawled") or "-", r.get("status") or "-",
                          r.get("indexability", "-"), r.get("indexability_status", "-"),
                          r.get("canonical_target") or "-", r.get("action", "")])
            widths = (60, 14, 16, 16, 16, 55, 30)

        for i, w in enumerate(widths, 1):
            col_letter = ws.cell(3, i).column_letter
            ws.column_dimensions[col_letter].width = w
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                cell.alignment = NO_WRAP
        ws.freeze_panes = "A4"

    wb.save(out_path)
    log(f"[DONE] {out_path}")


# Reasons whose default action text implicitly assumes the page still
# resolves 200 today ("add a canonical tag", "index this page", "improve
# internal linking", "remove the noindex tag") - GSC's recorded reason can be
# stale; if today's live check shows the page no longer even returns 200,
# that default action doesn't apply until the real, current problem
# (a 404/redirect/5xx) is dealt with first.
_ASSUMES_LIVE_200 = {
    "Discovered - currently not indexed",
    "Crawled - currently not indexed",
    "Soft 404",
    "Duplicate without user-selected canonical",
    "Excluded by 'noindex' tag",
}


def _derive_action(reason, info, status, indexability_status, url=None):
    """The Action column - grounded in the LIVE check just performed, not
    just the reason's static default text, so it never tells someone to fix
    something that's already resolved (or recommends an action that makes no
    sense for a page that doesn't even return 200 anymore)."""
    normalized = reason.translate(str.maketrans(_UNICODE_PUNCT_NORMALIZE))
    default = info["default_action"] or "Check manually"

    if status is None:
        return default

    # A static resource file (JS/CSS/image/font) isn't a content page - none
    # of the reason-specific "index it"/"add a canonical"/"remove noindex"
    # advice below is meaningful for one, regardless of which reason GSC
    # filed it under. See _is_static_asset()'s docstring for the real
    # example (a WordPress emoji script) that surfaced this.
    if url and _is_static_asset(url):
        return ("This is a static resource file (script/style/image/font), not a content page - "
                "it's not meant to be indexed and none of the usual indexing advice applies here.")

    # A live canonical pointing elsewhere is DELIBERATE and self-explanatory -
    # it's not "we forgot to index this", it's "this URL is explicitly
    # telling Google to index a different one instead". Confirmed live: GSC's
    # cached "Crawled - currently not indexed" reason can lag what the page
    # actually says today, producing the nonsensical "Need to index this
    # page" on a URL whose own canonical tag says otherwise. This check comes
    # BEFORE the reason-specific ones below since it's true regardless of
    # which reason GSC originally recorded.
    if indexability_status == "Canonicalised":
        return ("Already canonicalized to a different URL - that's a deliberate signal telling "
                "Google to index the canonical target instead, not a gap to fix by indexing this "
                "one. No action needed unless the canonical target itself is wrong.")

    if normalized == "Excluded by 'noindex' tag":
        if indexability_status == "noindex":
            return default
        if status == 200:
            return ("No longer noindex on today's live check - Search Console's record may be "
                    "stale. Verify with URL Inspection before assuming this still needs fixing.")
    elif indexability_status == "noindex":
        # Same "GSC's cached reason is stale" story, but for any OTHER reason
        # where the live check found a noindex tag GSC hasn't caught up to yet.
        return "This page now has a noindex tag on a live check - remove it if it should be indexed."

    if "Add a canonical tag" in default and status == 200:
        return "Add a canonical tag to the preferred version, if this page is useful enough to keep indexed."

    if normalized in _ASSUMES_LIVE_200 and status != 200:
        return (f"This URL now returns {status} live (not 200) - Search Console's record may be "
               f"stale. Verify and resolve the {status} first; \"{default}\" doesn't apply until it does.")

    return default


def _url_and_crawled(item):
    """urls entries can be a plain string (older callers, the CLI JSON path)
    or a {"url":, "last_crawled":} dict (gsc_audit's live table scrape,
    which reads the "Last crawled" column right off GSC's own drilldown
    table) - normalize either shape to (url, last_crawled_text_or_'')."""
    if isinstance(item, dict):
        return item.get("url"), item.get("last_crawled") or ""
    return item, ""


def _inspect_fallback(url, api_token, api_property_url):
    """Best-effort fallback via Search Console's OWN URL Inspection API,
    used only when a direct live HTTP check couldn't get any result at all
    (persistent network failure/block, not a real 429 - those are already
    retried directly in follow_redirect_chain()). Real per-property daily
    quota on Google's side, so this must stay a rare fallback, not a
    per-URL default. Returns None (silently) if unconfigured or the
    fallback call itself fails - the caller keeps the original "could not
    check" result either way, this only adds detail when it succeeds."""
    if not api_token or not api_property_url:
        return None
    try:
        import gsc_audit
        result = gsc_audit.inspect_url(api_token, api_property_url, url)
        idx = result.get("indexStatusResult") or {}
        return {"verdict": idx.get("verdict") or "", "coverage": idx.get("coverageState") or "",
                "last_crawl": idx.get("lastCrawlTime") or "", "robots_state": idx.get("robotsTxtState") or ""}
    except Exception as e:
        log(f"   [warn] URL Inspection fallback failed for {url}: {type(e).__name__}: {e}")
        return None


def process_reason(reason, urls, domain, sitemap_urls, homepage_url, stated_count=None,
                    max_workers=15, api_token=None, api_property_url=None):
    """Live-check every URL for one reason and build its row list - the
    shared per-reason pipeline main() and any other caller (e.g. a future
    web_app_batch.py route) should both use, so the live-check/redirect-
    suggestion logic never has to be duplicated.

    stated_count (optional): the reason's real affected-page count per GSC's
    own summary export (gsc_audit._read_summary_reasons) - when GSC's own
    per-reason export (capped at 1000 rows, same class of limit the SE
    Ranking PDF export already has to be flagged for elsewhere in this app)
    returned fewer URLs than that, a trailing note row says so explicitly
    rather than silently reporting a partial list as if it were complete.

    Each URL's live check is a network round trip (a live redirect-chain
    walk, or a single status check) with no dependency on any other URL's
    result - confirmed live these ran fully sequentially at ~1.2-1.3s/URL,
    turning a ~2,400-URL run into the better part of an hour. A small thread
    pool runs them concurrently instead; results are collected back into the
    ORIGINAL url order (not completion order) since the report's row order
    should match GSC's own list, not whichever finished first."""
    info = _reason_info(reason)
    redirect_relevant = reason.translate(str.maketrans(_UNICODE_PUNCT_NORMALIZE)) in (
        "Not found (404)", "Page with redirect")
    norm_urls = [_url_and_crawled(u) for u in urls]

    def _check_one(url, last_crawled):
        if redirect_relevant:
            return _build_redirect_row(url, reason, sitemap_urls, homepage_url, last_crawled=last_crawled)
        live = check_live_status(url)
        if live.get("status") is None:
            insp = _inspect_fallback(url, api_token, api_property_url)
            if insp:
                label = insp["coverage"] or insp["verdict"] or "no verdict"
                return {"kind": "simple", "url": url, "last_crawled": last_crawled, "status": None,
                        "indexability": "Unknown",
                        "indexability_status": f"Live check failed - Search Console's URL Inspection: {label}",
                        "canonical_target": "-",
                        "action": (f"Could not verify live directly; Search Console's own URL Inspection "
                                  f"says '{label}'" + (f" (last crawled {insp['last_crawl']})"
                                  if insp["last_crawl"] else "") + " - verify manually before acting.")}
        indexability, indexability_status, canonical_target = classify_indexability(url, live)
        action = _derive_action(reason, info, live.get("status"), indexability_status, url=url)
        return {"kind": "simple", "url": url, "last_crawled": last_crawled, "status": live.get("status"),
                "indexability": indexability, "indexability_status": indexability_status,
                "canonical_target": canonical_target or "-", "action": action}

    out = [None] * len(norm_urls)
    done = 0
    with ThreadPoolExecutor(max_workers=min(max_workers, max(1, len(norm_urls)))) as pool:
        futures = {pool.submit(_check_one, url, last_crawled): i
                  for i, (url, last_crawled) in enumerate(norm_urls)}
        for fut in as_completed(futures):
            i = futures[fut]
            url, last_crawled = norm_urls[i]
            try:
                out[i] = fut.result()
            except Exception as e:
                err = f"Check manually - live check failed ({type(e).__name__})"
                if redirect_relevant:
                    out[i] = {"kind": "redirect", "url": url, "last_crawled": last_crawled,
                              "current_status": "", "redirects_to": "", "destination_status": "",
                              "hop_count": "", "recommendation": err, "redirect_target": "-",
                              "target_status": "-"}
                else:
                    out[i] = {"kind": "simple", "url": url, "last_crawled": last_crawled, "status": None,
                              "indexability": "", "indexability_status": "", "canonical_target": "-",
                              "action": err}
            done += 1
            if done % 25 == 0 or done == len(norm_urls):
                log(f"   [{done}/{len(norm_urls)}] checked")

    if stated_count and len(norm_urls) < stated_count:
        remaining = stated_count - len(norm_urls)
        note = (f"+ {remaining} more page(s) affected by this issue - Search Console's own "
               f"export only returned {len(norm_urls)} of {stated_count} (GSC caps CSV exports "
               f"at 1,000 rows per reason). Use URL Inspection or a sitemap-based crawl for "
               f"the remaining pages.")
        if redirect_relevant:
            out.append({"kind": "redirect", "url": note, "last_crawled": "", "current_status": "",
                       "redirects_to": "", "destination_status": "", "hop_count": "",
                       "recommendation": "", "redirect_target": "", "target_status": ""})
        else:
            out.append({"kind": "simple", "url": note, "last_crawled": "", "status": None,
                       "indexability": "", "indexability_status": "", "canonical_target": "", "action": ""})
    return out


def _build_redirect_row(url, reason, sitemap_urls, homepage_url, last_crawled=""):
    """Full redirect-chain analysis for one URL, for the two reasons where
    it matters most (Not found (404), Page with redirect) - see
    follow_redirect_chain()'s docstring for why the whole chain, not just
    the final destination, is needed to tell a clean single-hop redirect
    apart from a chain, a loop, or one that dead-ends somewhere else.

    The fix is 3 separate columns rather than one text blob - "Recommendation"
    (what to do, short), "Suggested Redirect To" (the actual target URL, if
    one is being suggested), "Suggested Target Status" (that target's own
    live-confirmed status) - so the target URL is directly usable/copyable
    instead of buried in a sentence."""
    chain = follow_redirect_chain(url, max_hops=10)
    hop_count = chain["hop_count"]
    final_status = chain["final_status"]
    redirects_to = chain["chain"][1]["url"] if len(chain["chain"]) > 1 else "-"
    is_404_reason = reason.translate(str.maketrans(_UNICODE_PUNCT_NORMALIZE)) == "Not found (404)"

    if chain["loop"]:
        recommendation = "Redirect loop detected - unreachable. Fix immediately."
        redirect_target, target_status = "-", "-"
    elif final_status == 200 and hop_count <= 1:
        recommendation = "No action needed - already resolves cleanly to a live 200 page."
        redirect_target, target_status = "-", "-"
    elif final_status == 200 and hop_count > 1:
        recommendation = (f"Resolves live but via {hop_count} redirect hops - consolidate to a "
                          f"single 301 directly to →")
        redirect_target, target_status = chain["final_url"], "200 OK (confirmed live)"
    elif _is_static_asset(url):
        # A dead/erroring JS, CSS, image, or font file isn't a content page
        # competing for a ranking - it doesn't need a redirect to the
        # homepage (which would just be wrong/misleading) or any other
        # content page. If it's genuinely needed by the site, fix it where
        # it's referenced/hosted; if not, it can be safely ignored here.
        recommendation = ("This is a static resource file (script/style/image/font), not a "
                          "content page - redirecting it to another page wouldn't be meaningful. "
                          "If the site actually needs this file, fix where it's hosted directly; "
                          "otherwise no SEO action needed.")
        redirect_target, target_status = "-", "-"
    else:
        target = find_live_redirect_target(url, sitemap_urls, homepage_url)
        if target:
            recommendation = (f"Redirect the {'404' if is_404_reason else 'broken redirect'} page to →")
            redirect_target, target_status = target["final_url"], "200 OK (confirmed live)"
        else:
            recommendation = ("Could not confirm any live replacement page automatically (even the "
                              "homepage didn't resolve cleanly) - choose a redirect target manually.")
            redirect_target, target_status = "-", "-"

    return {"kind": "redirect", "url": url, "last_crawled": last_crawled,
           "current_status": chain["chain"][0]["status"] if chain["chain"] else None,
           "redirects_to": redirects_to, "destination_status": final_status, "hop_count": hop_count,
           "recommendation": recommendation, "redirect_target": redirect_target,
           "target_status": target_status}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--domain", required=True)
    ap.add_argument("--reason-urls", required=True,
                    help="JSON file: {reason_name: [url, ...], ...} - see gsc_audit.capture_index_coverage_urls()")
    ap.add_argument("--out", required=True)
    ap.add_argument("--brand", default=None)
    args = ap.parse_args()

    with open(args.reason_urls, "r", encoding="utf-8") as f:
        reason_urls = json.load(f)

    log(f"Discovering {args.domain}'s real sitemap for redirect suggestions...")
    sitemap_urls = georpt.get_sitemap_urls(args.domain, cap=2000)
    log(f"   {len(sitemap_urls)} sitemap URL(s) found.")
    homepage_url = f"https://{onpage2.safe_domain(args.domain)}/"

    onpage2.set_run_scale(sum(len(v) for v in reason_urls.values()))

    reason_rows = {}
    for i, (reason, urls) in enumerate(reason_urls.items(), 1):
        log(f"[{i}/{len(reason_urls)}] '{reason}' ({len(urls)} URL(s))...")
        reason_rows[reason] = process_reason(reason, urls, args.domain, sitemap_urls, homepage_url)

    build_report(args.domain, reason_rows, args.out, brand=args.brand)


if __name__ == "__main__":
    main()
