"""
generate_seranking_audit.py - Turns a SEranking Site Audit Excel export into the
polished, suggestion-filled final workbook the team currently builds by hand.

Input: a SEranking Site Audit .xlsx export (one sheet per issue type - JS/CSS
issues, broken links, missing/duplicate H1, title/description issues, etc).
Output: the SAME sheets (whatever is actually present - not a fixed 29), with:
  - Merged "Suggestion" cells forward-filled so every row carries its value.
  - Judgment columns (Suggested Title / Suggested H1 / Suggested Description /
    Broken link Suggestion) - REGENERATED from the page's own live content via
    the same title/description suggestion logic already used by the SEO On-Page
    report, so it's the same "don't touch what's already fine" behavior, not a
    blind rewrite. A page that's already optimized (or ranks well) gets left
    alone with a "No changes needed" note, exactly like the On-Page report.
  - Everything else (raw crawl-fact sheets: JS/CSS/image/nofollow-link issues)
    passed through UNCHANGED - that data and its generic advice text is already
    correct; this tool never invents new claims about it.

This never edits a live site - it only produces a report for human review.

Run:
    python generate_seranking_audit.py --in "Website Audit Report.xlsx" --out "Final Audit.xlsx" --brand "Acme"
"""
import os
import re
import sys
import zlib
import argparse
import urllib.request
import urllib.error
import urllib.parse
from pathlib import Path

ROOT = Path(__file__).parent.resolve()
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Reuse the exact same "is this actually optimized" + AI/heuristic suggestion
# logic the On-Page report uses, so behavior is consistent across both tools.
import generate_seo_onpage_phase2 as onpage2


def log(msg):
    # Windows' console can be a legacy codepage (cp1252) that can't encode
    # every character a scraped page title/URL might contain (e.g. Cyrillic)
    # - crashing on a log line about a harmless fetch failure took down the
    # entire run. Falls back to a lossy-but-safe encode rather than raising.
    try:
        print(msg, flush=True)
    except UnicodeEncodeError:
        enc = sys.stdout.encoding or "ascii"
        print(str(msg).encode(enc, errors="replace").decode(enc), flush=True)


# --------------------------------------------------------------------------- #
# Column-name based classification - works for "however many sheets" the input
# actually has, not a fixed list, since exports vary run to run.
# --------------------------------------------------------------------------- #
TITLE_COLS = {"suggested title", "suggestedtitle"}
DESC_COLS = {"suggested description"}
H1_COLS = {"suggested h1"}
# "page with isues" - a typo the team's own hand-built reports consistently
# use (verified against a real client-facing report) - matched here on input
# so those sheets aren't silently skipped, without ever writing the typo back
# out ourselves.
PAGE_COLS = ("page with issues", "page with isues", "page url", "target pages", "url")
EXISTING_TITLE_COLS = {"existing title", "existing title and h1 tag", "existing title & h1", "title"}
EXISTING_DESC_COLS = {"existing description", "description"}
EXISTING_H1_COLS = {"existing h1", "h1"}
ALT_COLS = {"suggested alt text", "alt text suggestion", "suggested image alt tag"}
CANON_COLS = {"suggested canonical", "recommended canonical tag", "recommended canonical"}
EXISTING_CANON_COLS = {"existing canonical", "existing canonical tag"}
IMAGE_URL_COLS = {"image url", "resource url"}

# Issue types where the team writes ONE generic developer note (not a
# per-page suggestion) in the first data row only - verified against a real
# client-facing report (Website Audit Report - anoasisofhealing.com.xlsx).
# Matched by sheet-name pattern; the boilerplate text is the team's own
# wording from that report, output under a clean "Suggestion For Developer"
# header regardless of how the input sheet happened to spell its own column.
GENERIC_ADVICE_SHEETS = [
    (re.compile(r"slow loading", re.I),
     "Optimize the HTML code for the pages specified in the report. This is important because if "
     "the HTML code of the page isn't optimized, the page will take longer to load. Also, consider "
     "checking your web server, as it may be the root of the problem. If optimizing your code "
     "doesn't help, consider moving to a faster web server."),
    (re.compile(r"orphaned page", re.I),
     "Review all orphaned pages in your sitemap.xml files and do either of the following: If a page "
     "is no longer needed, remove it; If a page has valuable content and brings traffic to your "
     "website, link to it from another page on your website; If a page serves a specific need and "
     "requires no internal linking, leave it as is."),
    (re.compile(r"crawl depth", re.I),
     "Make sure that pages with important content can be reached within a few clicks. If any of "
     "them are buried too deep in your site, consider changing your internal link architecture."),
    (re.compile(r"permanent redirect", re.I),
     "We've identified a permanent (301) redirect issue - please replace the URL in the source page "
     "with the final destination URL directly, instead of routing through a redirect."),
    (re.compile(r"text.{0,3}html.{0,3}ratio", re.I),
     "Split your webpage's text content and code into separate files and compare their size. If the "
     "size of your code file exceeds the size of the text file, review your page's HTML code and "
     "consider optimizing its structure and removing embedded scripts and styles."),
    (re.compile(r"unminified.*(javascript|css)|minif", re.I),
     "Minify your JavaScript and CSS files. If your webpage uses CSS and JS files that are hosted on "
     "an external site, contact the website owner and ask them to minify their files. If this issue "
     "doesn't affect your page load time, simply ignore it."),
]


def _norm(s):
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _find_col(headers, wanted):
    for i, h in enumerate(headers):
        if _norm(h) in wanted:
            return i
    return None


def _find_page_col(headers):
    for i, h in enumerate(headers):
        n = _norm(h)
        if any(n == p or n.startswith(p) for p in PAGE_COLS):
            return i
    return None


def _find_image_col(headers):
    return _find_col(headers, IMAGE_URL_COLS)


# --------------------------------------------------------------------------- #
# Merged-cell handling - a merged "Suggestion" cell only carries its value on
# the top-left anchor; every other cell in the range reads as None otherwise.
# --------------------------------------------------------------------------- #
def _forward_fill_merged(ws):
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row = merged_range.min_col, merged_range.min_row
        max_col, max_row = merged_range.max_col, merged_range.max_row
        top_left = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).value = top_left


def _read_sheet(ws):
    _forward_fill_merged(ws)
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v in (None, "") for v in row):
            continue
        rows.append(list(row))
    return headers, rows


def _apply_generic_advice(sheet_name, headers, rows):
    """Issue types the team never writes a per-page suggestion for (Slow
    loading speed, orphaned pages in sitemap, page crawl depth, permanent
    redirects, text/html ratio, unminified JS/CSS) - one boilerplate developer
    note in the FIRST data row only, blank everywhere else, matching the
    team's own reports. Note this is applied regardless of what the input
    sheet's own suggestion column already had (including a merge-forward-
    filled duplicate on every row) - it always collapses back to row 1 only."""
    text = None
    for pattern, boilerplate in GENERIC_ADVICE_SHEETS:
        if pattern.search(sheet_name):
            text = boilerplate
            break
    if text is None:
        return headers, rows

    sug_col = None
    for i, h in enumerate(headers):
        if "suggest" in _norm(h):
            sug_col = i
            break
    headers = list(headers)
    if sug_col is None:
        headers.append("Suggestion For Developer")
        sug_col = len(headers) - 1
    else:
        headers[sug_col] = "Suggestion For Developer"

    new_rows = []
    for i, row in enumerate(rows):
        row = list(row)
        while len(row) <= sug_col:
            row.append(None)
        row[sug_col] = text if i == 0 else None
        new_rows.append(row)
    return headers, new_rows


# --------------------------------------------------------------------------- #
# PDF input (e.g. a SEMrush Site Audit export) - stdlib-only text extraction
# (zlib, no new dependency to distribute to every machine). PDFs like this are
# typically a summary/overview rather than a per-URL table, so this pulls out
# readable text into its own sheet rather than attempting risky, likely-wrong
# structured-table extraction from what's usually prose + stat blocks.
# --------------------------------------------------------------------------- #
def _pdf_decompress_streams(raw):
    """Every FlateDecode-compressed stream in the PDF, decompressed. Best-effort -
    a PDF with other filters (rare for text-based exports) just yields fewer/no
    chunks rather than raising."""
    chunks = []
    for m in re.finditer(rb"<<(.*?)>>\s*stream\r?\n(.*?)endstream", raw, re.S):
        dict_txt, body = m.group(1), m.group(2)
        if b"FlateDecode" not in dict_txt:
            continue
        try:
            chunks.append(zlib.decompress(body))
        except Exception:
            continue
    return chunks


def _pdf_extract_text_from_content(content):
    """Pull the text-showing operators (Tj / TJ) out of one decompressed PDF
    content stream, resolving PDF's backslash string escapes."""
    out = []
    for m in re.finditer(rb"\((?:\\.|[^\\()])*\)", content):
        s = m.group(0)[1:-1]
        s = re.sub(rb"\\([()\\])", rb"\1", s)
        s = s.replace(b"\\n", b" ").replace(b"\\r", b" ")
        try:
            out.append(s.decode("latin-1"))
        except Exception:
            continue
    return out


_PDF_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _looks_like_real_text(s):
    """FlateDecode also compresses embedded font programs and other binary
    resources, not just page content - a '(...)' pulled out of one of those
    decodes to noise, not text. Require mostly-printable content before keeping
    a line, so that garbage never reaches the output sheet (openpyxl also
    outright rejects certain control characters)."""
    if not s or len(s) < 2:
        return False
    printable = sum(1 for c in s if c == " " or (32 <= ord(c) < 127) or ord(c) > 160)
    return (printable / len(s)) > 0.85


def _ensure_pdfminer():
    """pdfminer.six properly decodes embedded custom font glyph encodings that
    the lightweight stdlib extractor below can't - confirmed against a real
    SEranking PDF export where the stdlib path recovered ~0 real words and
    pdfminer recovered the full report cleanly. Installed on first use into
    this same embedded interpreter (pure top-level import check, no restart
    needed); if pip/PyPI isn't reachable (offline machine, blocked network),
    this just returns None and the caller falls back to the stdlib extractor."""
    try:
        from pdfminer.high_level import extract_text
        return extract_text
    except ImportError:
        pass
    try:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "pdfminer.six", "--quiet"],
                        timeout=120, capture_output=True, check=True)
        from pdfminer.high_level import extract_text
        return extract_text
    except Exception:
        return None


def extract_pdf_text(path):
    """Return the PDF's visible text as a list of lines. Tries pdfminer.six
    first (handles real-world custom-font-encoded PDFs correctly); falls back
    to a lightweight stdlib-only extraction (works for PDFs with literal text
    strings, e.g. simpler/older exports) if pdfminer is unavailable or fails."""
    extract_text = _ensure_pdfminer()
    if extract_text is not None:
        try:
            import logging
            logging.getLogger("pdfminer").setLevel(logging.ERROR)
            text = extract_text(path)
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            if lines:
                return lines
        except Exception:
            pass

    with open(path, "rb") as f:
        raw = f.read()
    lines = []
    for stream in _pdf_decompress_streams(raw):
        pieces = _pdf_extract_text_from_content(stream)
        if pieces:
            line = _PDF_CONTROL_CHARS.sub("", "".join(pieces)).strip()
            if line and _looks_like_real_text(line):
                lines.append(line)
    return lines


def build_pdf_summary_rows(pdf_path):
    """[[line_number, text], ...] ready to drop straight into a sheet via
    _write_sheet - the PDF's content preserved for human review, since a
    summary-style PDF rarely has enough per-URL structure to safely auto-suggest
    from without risking wrong/invented claims.

    extract_pdf_text() tries pdfminer.six first, which correctly decodes the
    embedded custom font glyph encodings some PDFs (confirmed with real
    SEranking/SEMrush exports) use instead of literal text strings - the
    lightweight stdlib fallback can only recover font metadata noise from
    those, never the actual words. The quality check below is a last-resort
    safety net for when even pdfminer can't get real text (e.g. a genuinely
    scanned/image-only PDF, or pdfminer/PyPI unavailable on this machine) -
    showing that noise as if it were real content would be actively
    misleading, so it refuses to return it and reports the limitation
    honestly instead."""
    lines = extract_pdf_text(pdf_path)
    words = " ".join(lines).split()
    has_real_words = sum(1 for w in words if w.isalpha() and len(w) > 2)
    if not lines or has_real_words < 5:
        return [[1, "This PDF's text could not be reliably extracted - it likely renders "
                    "text through an embedded custom font rather than literal text (common "
                    "in SEMrush/design-tool exports), and pdfminer.six could not be installed "
                    "or failed to decode it either. Please review the PDF manually, or "
                    "provide the underlying data as an .xlsx export instead."]]
    return [[i + 1, line] for i, line in enumerate(lines)]


# --------------------------------------------------------------------------- #
# Live page fetch - best-effort, never raises. Same shape suggest_meta() wants.
# --------------------------------------------------------------------------- #
_UA = "Mozilla/5.0 SERankingAuditBot"


def _safe_url(url):
    """Non-ASCII URLs (e.g. a Cyrillic path segment) crash deep inside
    http.client, which builds the raw HTTP request line as ASCII - percent-
    encode the path/query the same way a browser would before requesting."""
    try:
        parts = urllib.parse.urlsplit(url)
        path = urllib.parse.quote(parts.path, safe="/%")
        query = urllib.parse.quote(parts.query, safe="=&%")
        return urllib.parse.urlunsplit((parts.scheme, parts.netloc, path, query, parts.fragment))
    except Exception:
        return url


def _fetch_page_data(url):
    try:
        req = urllib.request.Request(_safe_url(url), headers={"User-Agent": _UA})
        with urllib.request.urlopen(req, timeout=15) as r:
            html = r.read().decode("utf-8", "ignore")
    except Exception as e:
        log(f"   [warn] could not fetch {url}: {type(e).__name__}: {e}")
        html = ""
    return onpage2._parse_html(html, url, 200 if html else 0) if html else {
        "url": url, "title": onpage2.MISSING, "description": onpage2.MISSING,
        "h1": onpage2.MISSING, "body_text": "",
    }


def _guess_keywords(page_url):
    """No keyword sheet is guaranteed to exist for this input - fall back to the
    page's own slug words as a rough target keyword, same idea as auto-discovery
    in the On-Page report when no keyword list is supplied."""
    path = re.sub(r"^https?://[^/]+/?", "", page_url).strip("/")
    words = re.split(r"[-_/]+", path)
    words = [w for w in words if w and not w.isdigit()]
    return [" ".join(words[-4:])] if words else []


def _title_relevant_to_h1(title, h1):
    """A content-relevance check using the page's OWN H1 - a far more reliable
    signal of what the page is actually about than a URL-slug guess. Confirmed
    real false positive: aceabseiling.com.au/caulking-sealing-water-leaking/'s
    existing title "Rope Access Water Leak & Sealing Solutions Sydney NSW" was
    flagged and rewritten purely because it doesn't literally say "caulking"
    (the URL-guessed keyword), even though the title is clearly about the same
    topic the H1 describes. Only used to veto a suggestion that would otherwise
    fire SOLELY on the URL-guessed-keyword check - a missing/generic/too-short
    title is still always suggested regardless."""
    if not h1 or h1 == onpage2.MISSING or not title or title == onpage2.MISSING:
        return False
    h1_words = {w for w in re.findall(r"[a-z0-9]+", h1.lower()) if len(w) > 3}
    if not h1_words:
        return False
    title_words = {w for w in re.findall(r"[a-z0-9]+", title.lower()) if len(w) > 3}
    overlap = h1_words & title_words
    return len(overlap) >= max(1, len(h1_words) // 3)


_brand_cache = {}


def _suggest_for_page(page_url, brand):
    if page_url in _brand_cache:
        return _brand_cache[page_url]
    pd = _fetch_page_data(page_url)
    keywords = _guess_keywords(page_url)
    result = onpage2.suggest_meta(pd, keywords, brand)
    # The "keyword" fed into suggest_meta above is only GUESSED from the URL
    # path, not a real target keyword - judging "does the existing title need
    # replacing" by that alone produces false positives for perfectly good
    # titles phrased differently than the URL. If the title itself is fine
    # (not missing/generic/too-short - suggest_meta's other checks already
    # cover those) and it's topically related to the page's own H1, leave it
    # alone instead of rewriting it.
    existing_title = pd.get("title")
    if (result.get("suggested_title") and result["suggested_title"] != "No changes needed - existing tag is already optimized"
            and existing_title and existing_title != onpage2.MISSING and len(existing_title.strip()) >= 15
            and existing_title.strip().lower() not in onpage2._GENERIC_TITLES
            and _title_relevant_to_h1(existing_title, pd.get("h1"))):
        result = dict(result)
        result["suggested_title"] = "No changes needed - title already reflects this page's content."
    _brand_cache[page_url] = result
    return result


def _suggest_alt_text(page_url, img_src, brand):
    """Real AI-generated ALT text for one image - same approach GEO's Technical
    Optimization report already uses (generate_geo_report._suggest_alt_text),
    ported here so SEranking's "Alt text missing" sheet gets a real per-image
    suggestion instead of just passing the raw HTTP-status columns through."""
    page_title = ""
    try:
        s = _suggest_for_page(page_url, brand)
        et = s.get("existing_title")
        page_title = et if et and et != onpage2.MISSING else ""
    except Exception:
        pass
    prompt = (
        f"Write a concise, descriptive image ALT text (under 125 characters) for an image on "
        f"{brand}'s page at {page_url} (page topic: {page_title or 'unknown'}). "
        f"The image file is: {img_src}. Infer what the image likely shows from the filename/page "
        "topic. Return ONLY the ALT text, no quotes, no extra commentary."
    )
    result = onpage2._ai_suggest(f'Return ONLY JSON: {{"alt": "..."}}\n\n{prompt}')
    if isinstance(result, dict) and result.get("alt"):
        return str(result["alt"]).strip()
    # Heuristic fallback - filename-derived, same as GEO's version
    name = img_src.rsplit("/", 1)[-1].rsplit(".", 1)[0]
    name = re.sub(r"[-_]+", " ", name).strip() or "image"
    return f"{name} - {brand}".strip(" -")


_canon_cache = {}


def _suggest_canonical(page_url):
    """Real canonical-tag recommendation via the same logic the On-Page report
    uses (onpage2.recommend_canonical), for SEranking's "Canonical tag" sheet -
    that sheet's template already ships an "Existing"/"Recommended Canonical
    tag" column pair, just never filled in before."""
    if page_url in _canon_cache:
        return _canon_cache[page_url]
    pd = _fetch_page_data(page_url)
    result = onpage2.recommend_canonical(pd)
    _canon_cache[page_url] = result
    return result


# --------------------------------------------------------------------------- #
# Broken-link judgment - a dead link either has a discoverable redirect target
# (suggest that) or genuinely doesn't (suggest removal). Never invents a URL.
# --------------------------------------------------------------------------- #
def _suggest_broken_link_fix(url):
    try:
        req = urllib.request.Request(_safe_url(url), headers={"User-Agent": _UA}, method="HEAD")
        with urllib.request.urlopen(req, timeout=12) as r:
            final = r.geturl()
            if final and final.rstrip("/") != url.rstrip("/"):
                return f"Redirects to {final} - update the link to point there directly."
            return "No changes needed - link is live."
    except urllib.error.HTTPError as e:
        if e.code in (404, 410):
            return "Remove This URL"
        return f"HTTP {e.code} - please verify manually."
    except Exception:
        return "Could not verify automatically - please check manually."


# --------------------------------------------------------------------------- #
# Per-sheet suggestion filling
# --------------------------------------------------------------------------- #
# Sheet-NAME based fallback for inputs that don't already carry pre-built
# "Suggested X" columns (a plain issue-list export, not SEranking's own template
# with those columns already present but blank) - if the sheet is clearly about
# one of these issue types and has a page/URL column, the right suggestion
# column is added rather than requiring it to already exist.
_NAME_JUDGMENT_RULES = [
    (re.compile(r"h1", re.I), "h1"),
    (re.compile(r"title", re.I), "title"),
    (re.compile(r"description", re.I), "desc"),
    (re.compile(r"4xx|broken|dead.?link", re.I), "broken"),
    (re.compile(r"alt.?text|alt.?tag|missing.?alt", re.I), "alt"),
    (re.compile(r"canonical", re.I), "canon"),
]


def _apply_suggestions(sheet_name, headers, rows, brand):
    page_col = _find_page_col(headers)
    title_col = _find_col(headers, TITLE_COLS)
    desc_col = _find_col(headers, DESC_COLS)
    h1_col = _find_col(headers, H1_COLS)
    broken_col = _find_col(headers, {"broken link suggestion"})
    existing_title_col = _find_col(headers, EXISTING_TITLE_COLS)
    existing_desc_col = _find_col(headers, EXISTING_DESC_COLS)
    existing_h1_col = _find_col(headers, EXISTING_H1_COLS)
    alt_col = _find_col(headers, ALT_COLS)
    canon_col = _find_col(headers, CANON_COLS)
    existing_canon_col = _find_col(headers, EXISTING_CANON_COLS)
    image_col = _find_image_col(headers)

    has_any_suggestion_col = any(c is not None for c in (title_col, desc_col, h1_col, broken_col, alt_col, canon_col))
    if page_col is not None and not has_any_suggestion_col:
        # No pre-built suggestion column - infer what's needed from the sheet name
        # and add it, so a plain (non-SEranking-template) issue list still gets
        # real suggestions instead of silently passing through untouched.
        kinds = {kind for pattern, kind in _NAME_JUDGMENT_RULES if pattern.search(sheet_name)}
        if kinds:
            if "title" in kinds:
                headers = headers + ["Suggested Title"]; title_col = len(headers) - 1
            if "desc" in kinds:
                headers = headers + ["Suggested Description"]; desc_col = len(headers) - 1
            if "h1" in kinds:
                headers = headers + ["Suggested H1"]; h1_col = len(headers) - 1
            if "broken" in kinds:
                headers = headers + ["Broken link Suggestion"]; broken_col = len(headers) - 1
            if "alt" in kinds and image_col is not None:
                headers = headers + ["Suggested Alt Text"]; alt_col = len(headers) - 1
            if "canon" in kinds:
                headers = headers + ["Recommended Canonical Tag"]; canon_col = len(headers) - 1
            rows = [list(r) + [None] * (len(headers) - len(r)) for r in rows]

    if page_col is None or not (title_col is not None or desc_col is not None
                                 or h1_col is not None or broken_col is not None
                                 or existing_title_col is not None or existing_desc_col is not None
                                 or existing_h1_col is not None or alt_col is not None
                                 or canon_col is not None or existing_canon_col is not None):
        return headers, rows  # not a judgment sheet - pass through unchanged

    log(f"   Generating suggestions for '{sheet_name}' ({len(rows)} row(s))...")
    out = []
    for row in rows:
        row = list(row)
        page_url = str(row[page_col] or "").strip() if page_col < len(row) else ""
        if page_url and page_url.startswith("http"):
            if broken_col is not None:
                row[broken_col] = _suggest_broken_link_fix(page_url)
            if alt_col is not None:
                img_url = str(row[image_col] or "").strip() if (image_col is not None and image_col < len(row)) else ""
                row[alt_col] = _suggest_alt_text(page_url, img_url, brand) if img_url else \
                    "Could not determine the image URL - please check manually."
            if canon_col is not None or existing_canon_col is not None:
                canon = _suggest_canonical(page_url)
                if canon_col is not None:
                    row[canon_col] = canon["recommended"]
                if existing_canon_col is not None:
                    row[existing_canon_col] = canon["existing"] if canon["existing"] != onpage2.MISSING else "Missing"
            if (title_col is not None or desc_col is not None or h1_col is not None
                    or existing_title_col is not None or existing_desc_col is not None
                    or existing_h1_col is not None):
                s = _suggest_for_page(page_url, brand)
                if title_col is not None:
                    row[title_col] = s["suggested_title"]
                if desc_col is not None:
                    row[desc_col] = s["suggested_description"]
                if h1_col is not None:
                    row[h1_col] = s["suggested_h1"]
                # Fill with the REAL, freshly-checked existing value from the live
                # page - never leave it blank just because the uploaded sheet's own
                # export happened to have it blank (that export can be stale, or
                # SEranking's own crawler can miss a tag a direct fetch picks up).
                # A genuinely missing tag is written as "Missing" explicitly, not
                # left blank, so a blank cell always means "not yet checked" and
                # never "confirmed absent" - the two are meaningfully different for
                # the team's review.
                if existing_title_col is not None:
                    row[existing_title_col] = s["existing_title"] if s["existing_title"] != onpage2.MISSING else "Missing"
                if existing_desc_col is not None:
                    row[existing_desc_col] = s["existing_description"] if s["existing_description"] != onpage2.MISSING else "Missing"
                if existing_h1_col is not None:
                    row[existing_h1_col] = s["existing_h1"] if s["existing_h1"] != onpage2.MISSING else "Missing"
        out.append(row)
    return headers, out


# --------------------------------------------------------------------------- #
# Output workbook - house style: navy header, no-wrap body (narrow columns,
# long text overflows into the next empty cell - Excel/Sheets' normal
# behavior - rather than every row growing tall to fit wrapped text).
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NO_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _write_sheet(wb_out, name, headers, rows):
    safe_name = re.sub(r"[\[\]:\\/?*]", "_", name)[:31] or "Sheet"
    base = safe_name
    n = 1
    while safe_name in wb_out.sheetnames:
        n += 1
        safe_name = f"{base[:28]}_{n}"
    ws = wb_out.create_sheet(safe_name)
    ws.append(headers)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = NO_WRAP
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = NO_WRAP
    for i, h in enumerate(headers, 1):
        col_letter = ws.cell(1, i).column_letter
        ws.column_dimensions[col_letter].width = min(28, max(12, len(str(h)) + 2))
    ws.freeze_panes = "A2"


# --------------------------------------------------------------------------- #
# Zip-of-.xls input (SEranking's "Pages" export downloaded per-issue, zipped
# together by the team instead of the single combined .xlsx). Each file is a
# legacy binary .xls (Composite Document / BIFF format, not .xlsx) named just
# "pages_<domain>_<timestamp>.xls" - no issue name in the filename or inside
# the file, so which issue each one covers has to be inferred from its column
# set instead of a sheet name (the way the combined-.xlsx path infers it).
# --------------------------------------------------------------------------- #
def _ensure_xlrd():
    """xlrd is the only maintained reader for legacy binary .xls (openpyxl only
    reads .xlsx). Installed on first use into this same embedded interpreter,
    same pattern as _ensure_pdfminer()."""
    try:
        import xlrd
        return xlrd
    except ImportError:
        pass
    try:
        import subprocess
        subprocess.run([sys.executable, "-m", "pip", "install", "xlrd==2.0.1", "--quiet"],
                        timeout=120, capture_output=True, check=True)
        import xlrd
        return xlrd
    except Exception:
        return None


def _read_seranking_zip(zip_path):
    """Returns [(filename, headers, rows), ...] for every .xls/.xlsx member.
    ignore_workbook_corruption=True is required - confirmed against a real
    SEranking zip export where 3 of 5 files raised a CompDocError without it
    despite being perfectly readable files (a known false-positive in xlrd's
    strict OLE2 stream validation, not actual corruption)."""
    import zipfile
    xlrd = _ensure_xlrd()
    out = []
    with zipfile.ZipFile(zip_path) as zf:
        names = [n for n in zf.namelist() if n.lower().endswith((".xls", ".xlsx")) and not n.startswith("__MACOSX")]
        for name in names:
            data = zf.read(name)
            try:
                if name.lower().endswith(".xlsx"):
                    import io
                    wb_in = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
                    ws_in = wb_in[wb_in.sheetnames[0]]
                    headers, rows = _read_sheet(ws_in)
                elif xlrd is not None:
                    wb_in = xlrd.open_workbook(file_contents=data, ignore_workbook_corruption=True)
                    sh = wb_in.sheet_by_index(0)
                    headers = [str(v).strip() if v is not None else "" for v in sh.row_values(0)]
                    rows = [list(sh.row_values(r)) for r in range(1, sh.nrows)
                            if any(v not in (None, "") for v in sh.row_values(r))]
                else:
                    log(f"   [warn] xlrd unavailable - skipping '{name}'")
                    continue
            except Exception as e:
                log(f"   [warn] could not read '{name}': {type(e).__name__}: {e}")
                continue
            if headers:
                out.append((name, headers, rows))
    return out


_GENERIC_XLS_STEMS = {"report", "export", "data", "sheet", "untitled", "download"}


def _display_name_from_filename(fname):
    """The zip-of-per-issue-exports format names each file exactly after the
    issue it covers (e.g. "Duplicate H1.xls", "Alt text missing.xls") - a far
    more reliable signal than guessing from columns, since SEranking's own
    per-issue exports often DON'T include a column literally saying "Duplicate
    H1" or "Alt text missing" (confirmed real case: Duplicate H1.xls's actual
    columns are just Page URL/H1/H1 length/... - no "duplicate" column at all,
    so the old column-only classifier mislabeled it as generic "Word counts
    Issues" and it silently got no H1 suggestions, unlike its sibling H1
    sheets). Returns None (caller falls back to column-based guessing) for a
    generic/uninformative filename."""
    stem = re.sub(r"\.(xlsx?|csv)$", "", fname, flags=re.I).strip()
    if not stem or _norm(stem) in _GENERIC_XLS_STEMS:
        return None
    return stem


# Column-signature -> display name. Fallback for when the filename itself
# isn't usable (see _display_name_from_filename above) - matched against a
# normalized set of header names; order matters (more specific signatures
# checked first). The chosen name is fed into the SAME
# _apply_suggestions()/_NAME_JUDGMENT_RULES path the combined-.xlsx sheets
# use, so e.g. "Title Issues" still gets a real "Suggested Title" column
# generated - no separate suggestion logic needed.
def _classify_xls_report(headers):
    cols = {_norm(h) for h in headers}
    has = lambda *names: all(n in cols for n in names)
    if has("title", "duplicate title"):
        return "Title Issues"
    if has("description", "duplicate description"):
        return "Description Issues"
    if has("h1", "duplicate h1"):
        return "H1 Issues"
    if "target url" in cols and "target url http status code" in cols:
        return "Redirect Chains"
    if any(n in cols for n in ("blocked by robots.txt", "robots meta tag", "x-robots-tag")):
        return "Robots Blocking Issues"
    if "page http status code" in cols:
        return "4XX HTTP Status Codes"
    # Unrecognized column set - name it from whatever non-generic column it has,
    # so it's still identifiable in the output rather than a bare "Sheet".
    distinctive = [h for h in headers if _norm(h) not in
                   ("page url", "referring pages", "is in sitemap", "url length")]
    return (distinctive[0] if distinctive else "Pages") + " Issues"


def _write_all_issue_sheet(wb_out, issue_sheet_names, label):
    """"All issue" index sheet (verified against a real client-facing report) -
    lists every issue-type sheet actually present in this run with a "Click
    Here For Solution" pointer to it. The team's own version also shows an
    overall "Issues : NN%" score, but that comes from SEranking's own site
    health scoring model which this script has no access to (it only ever
    sees the per-issue exports, never an aggregate score) - shown as "Issues
    Found" instead of guessing a number, consistent with never fabricating
    data this script can't actually verify."""
    ws = wb_out.create_sheet("All issue")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(1, 1, "Issues Found").font = header_font
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).alignment = left
    ws.cell(1, 2, label).font = header_font
    ws.cell(1, 2).fill = header_fill
    ws.cell(1, 2).alignment = left
    for r, name in enumerate(issue_sheet_names, 2):
        ws.cell(r, 1, name).alignment = left
        ws.cell(r, 2, "Click Here For Solution").alignment = left
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 34
    wb_out.move_sheet("All issue", offset=-len(wb_out.sheetnames) + 1)


def process_workbook(in_path, out_path, brand, pdf_path=None, zip_path=None):
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    issue_sheet_names = []

    # Pre-scan total row count across all sources so the AI paid-tier gate
    # (onpage2.BULK_AI_PAGE_LIMIT) reflects the real scale of this run before
    # any suggestion generation starts - SEranking sheets are often huge
    # (e.g. 800+ rows for a single issue export).
    total_rows = 0
    if in_path:
        wb_scan = openpyxl.load_workbook(in_path, data_only=True, read_only=True)
        for sheet_name in wb_scan.sheetnames:
            total_rows += max(0, wb_scan[sheet_name].max_row - 1)
        wb_scan.close()
    if zip_path:
        for fname, headers, rows in _read_seranking_zip(zip_path):
            total_rows += len(rows)
    onpage2.set_run_scale(total_rows)

    if in_path:
        log(f"Reading: {in_path}")
        wb_in = openpyxl.load_workbook(in_path, data_only=True)
        for i, sheet_name in enumerate(wb_in.sheetnames):
            log(f"[{i + 1}/{len(wb_in.sheetnames)}] Processing '{sheet_name}'...")
            # An "All issue"-style summary sheet already in the input (e.g. a
            # previously-generated report accidentally fed back in) would
            # collide with the one this script builds below - skip copying
            # it through rather than producing a workbook with two sheets of
            # the same name (undefined/broken behavior in Excel).
            if _norm(sheet_name) == "all issue":
                continue
            ws_in = wb_in[sheet_name]
            headers, rows = _read_sheet(ws_in)
            if not headers:
                continue
            headers, rows = _apply_suggestions(sheet_name, headers, rows, brand)
            headers, rows = _apply_generic_advice(sheet_name, headers, rows)
            _write_sheet(wb_out, sheet_name, headers, rows)
            issue_sheet_names.append(sheet_name)

    if zip_path:
        log(f"Reading zip: {zip_path}")
        members = _read_seranking_zip(zip_path)
        for i, (fname, headers, rows) in enumerate(members):
            display_name = _display_name_from_filename(fname) or _classify_xls_report(headers)
            log(f"[{i + 1}/{len(members)}] '{fname}' -> '{display_name}' ({len(rows)} row(s))...")
            if not headers:
                continue
            if _norm(display_name) == "all issue":
                continue
            headers, rows = _apply_suggestions(display_name, headers, rows, brand)
            headers, rows = _apply_generic_advice(display_name, headers, rows)
            _write_sheet(wb_out, display_name, headers, rows)
            issue_sheet_names.append(display_name)

    if pdf_path:
        log(f"Reading PDF: {pdf_path}")
        rows = build_pdf_summary_rows(pdf_path)
        log(f"   Extracted {len(rows)} line(s) of text.")
        _write_sheet(wb_out, "PDF Summary", ["#", "Text"], rows)

    if issue_sheet_names:
        label = brand or Path(in_path or zip_path or out_path).stem
        _write_all_issue_sheet(wb_out, issue_sheet_names, label)

    wb_out.save(out_path)
    log(f"[DONE] {out_path}")


# --------------------------------------------------------------------------- #
# Site Audit report - the SAME polished output (styling via _write_sheet(),
# suggestions via onpage2.suggest_meta() / _suggest_alt_text()) as
# process_workbook() above, but sourced from site_audit.py's already-computed
# results dict instead of a parsed SE Ranking .xlsx export. No xlsx-parsing
# step in between - that would just be a synthetic intermediate to throw away.
# --------------------------------------------------------------------------- #
def _sa_page_lookup(results):
    return {p.get("url"): p for p in (results.get("pages") or []) if p.get("url")}


def _sa_page_data_for(url, page_lookup):
    """A minimal onpage2-shaped page_data dict (title/description/h1/url) for
    suggest_meta() - built from Site Audit's own already-fetched per-page data
    (site_audit._extract_page_data), never a second live fetch of the same
    page. "body_text" isn't collected by Site Audit's lightweight extractor
    (it only tracks a word count, not the text itself), so it's left blank -
    suggest_meta() already treats that as "(not available)" and falls back to
    the page's H1/URL-slug for its heuristic and AI-prompt topic, exactly like
    a page with no visible body text would in the existing tool."""
    p = page_lookup.get(url) or {}
    title = p.get("title") or onpage2.MISSING
    desc = p.get("description") or onpage2.MISSING
    h1s = p.get("h1s") or []
    h1 = h1s[0] if h1s else onpage2.MISSING
    return {"url": url, "title": title, "description": desc, "h1": h1, "body_text": ""}


def _sa_suggest(url, brand, page_lookup, cache):
    if url in cache:
        return cache[url]
    result = onpage2.suggest_meta(_sa_page_data_for(url, page_lookup), [], brand)
    cache[url] = result
    return result


def _sa_broken_link_suggestion(item):
    """Reuses health_audit.check_broken_links' own suggested_redirect (already
    computed during the audit's crawl - never regenerated here) and mirrors
    _suggest_broken_link_fix()'s exact wording above for the no-redirect cases,
    so the phrasing is identical whether it came from a live HEAD check
    (SEranking-export path) or Site Audit's own crawl (this path)."""
    redirect = item.get("suggested_redirect")
    if redirect:
        return f"Redirects to {redirect} - update the link to point there directly."
    code = item.get("code")
    if code in (404, 410):
        return "Remove This URL"
    if code:
        return f"HTTP {code} - please verify manually."
    return "Could not verify automatically - please check manually."


def _sa_broken_links_sheet(items, page_lookup, brand, cache):
    headers = ["Page URL", "Status Code", "Found On", "Broken link Suggestion"]
    rows = [[it.get("url", ""), it.get("code", ""), "; ".join(it.get("found_on") or []),
             _sa_broken_link_suggestion(it)] for it in items]
    return headers, rows


def _sa_missing_titles_sheet(items, page_lookup, brand, cache):
    headers = ["Page URL", "Existing Title", "Suggested Title"]
    rows = []
    for it in items:
        url = it.get("url", "")
        s = _sa_suggest(url, brand, page_lookup, cache)
        rows.append([url, "Missing", s["suggested_title"]])
    return headers, rows


def _sa_duplicate_titles_sheet(items, page_lookup, brand, cache):
    headers = ["Title", "Page URL", "Suggested Title"]
    rows = []
    for it in items:
        title = it.get("title", "")
        for url in it.get("urls") or []:
            s = _sa_suggest(url, brand, page_lookup, cache)
            rows.append([title, url, s["suggested_title"]])
    return headers, rows


def _sa_missing_meta_sheet(items, page_lookup, brand, cache):
    headers = ["Page URL", "Existing Description", "Suggested Description"]
    rows = []
    for it in items:
        url = it.get("url", "")
        s = _sa_suggest(url, brand, page_lookup, cache)
        rows.append([url, "Missing", s["suggested_description"]])
    return headers, rows


def _sa_duplicate_meta_sheet(items, page_lookup, brand, cache):
    headers = ["Description", "Page URL", "Suggested Description"]
    rows = []
    for it in items:
        desc = it.get("description", "")
        for url in it.get("urls") or []:
            s = _sa_suggest(url, brand, page_lookup, cache)
            rows.append([desc, url, s["suggested_description"]])
    return headers, rows


def _sa_missing_h1_sheet(items, page_lookup, brand, cache):
    headers = ["Page URL", "Existing H1", "Suggested H1"]
    rows = []
    for it in items:
        url = it.get("url", "")
        s = _sa_suggest(url, brand, page_lookup, cache)
        rows.append([url, "Missing", s["suggested_h1"]])
    return headers, rows


def _sa_multiple_h1_sheet(items, page_lookup, brand, cache):
    headers = ["Page URL", "H1 Count", "Suggested H1"]
    rows = []
    for it in items:
        url = it.get("url", "")
        s = _sa_suggest(url, brand, page_lookup, cache)
        rows.append([url, it.get("count", ""), s["suggested_h1"]])
    return headers, rows


def _sa_missing_alt_sheet(items, page_lookup, brand, cache):
    headers = ["Page URL", "Image URL", "Suggested Alt Text"]
    rows = []
    for it in items:
        url = it.get("url", "")
        page = page_lookup.get(url) or {}
        srcs = page.get("missing_alt_srcs") or []
        if srcs:
            for src in srcs:
                rows.append([url, src, _suggest_alt_text(url, src, brand)])
        else:
            # Older cached results (from before Site Audit tracked per-image
            # srcs) or a page whose "pages" entry didn't survive - fall back to
            # the page-level count rather than guessing at an image URL.
            rows.append([url, "", f"{it.get('count', 0)} image(s) missing alt text on this page - "
                                   "add descriptive alt text for each manually."])
    return headers, rows


def _sa_thin_content_sheet(items, page_lookup, brand, cache):
    headers = ["Page URL", "Word Count", "Suggestion"]
    rows = [[it.get("url", ""), it.get("word_count", ""),
             "Below the recommended minimum word count for this page type - expand it with more "
             "unique, useful content that covers the topic in depth."] for it in items]
    return headers, rows


def _sa_redirect_chains_sheet(items, page_lookup, brand, cache):
    headers = ["Page URL", "Hops", "Final URL", "Suggestion"]
    rows = [[it.get("url", ""), it.get("hops", ""), it.get("final_url", ""),
             "Update the original link(s) to point directly to the final destination URL, removing "
             "the intermediate redirect hop(s)."] for it in items]
    return headers, rows


def _sa_noindex_sheet(items, page_lookup, brand, cache):
    headers = ["Page URL", "Robots Meta", "Suggestion"]
    rows = [[it.get("url", ""), it.get("robots", ""),
             "This page is set to noindex - confirm that's intentional; if not, remove the noindex "
             "directive so the page can be indexed."] for it in items]
    return headers, rows


def _sa_summary_only_sheet(items, page_lookup, brand, cache):
    headers = ["Summary"]
    return headers, [[it.get("summary", "")] for it in items]


# (sheet display name, issues-dict key, row-builder). Order controls sheet
# order in the output workbook. Only issue types Site Audit actually reports
# something equivalent to an SE Ranking Audit sheet for are included here -
# see build_report_from_site_audit()'s docstring for the ones that aren't.
_SA_SHEET_SPECS = [
    ("Broken Links", "broken_links", _sa_broken_links_sheet),
    ("Missing Titles", "missing_titles", _sa_missing_titles_sheet),
    ("Duplicate Titles", "duplicate_titles", _sa_duplicate_titles_sheet),
    ("Missing Meta Descriptions", "missing_meta", _sa_missing_meta_sheet),
    ("Duplicate Meta Descriptions", "duplicate_meta", _sa_duplicate_meta_sheet),
    ("Missing H1", "missing_h1", _sa_missing_h1_sheet),
    ("Multiple H1 Tags", "multiple_h1", _sa_multiple_h1_sheet),
    ("Missing Alt Text", "missing_alt_images", _sa_missing_alt_sheet),
    ("Thin Content Pages", "thin_content_pages", _sa_thin_content_sheet),
    ("Redirect Chains", "redirect_chains", _sa_redirect_chains_sheet),
    ("Noindex Pages", "noindex_pages", _sa_noindex_sheet),
    ("Sitemap Issues", "sitemap_issues", _sa_summary_only_sheet),
    ("Robots.txt Issues", "robots_issues", _sa_summary_only_sheet),
]


def build_report_from_site_audit(results, out_path, brand=None):
    """Site Audit's results dict (site_audit.run_site_audit()'s return value) ->
    the same polished, suggestion-filled workbook process_workbook() builds
    from a real SE Ranking export - one sheet per issue type actually present
    (not a fixed list), navy header row, "All issue" index sheet, suggestions
    generated by the exact same onpage2.suggest_meta()/_suggest_alt_text()
    logic that path uses.

    Limitations (flagged rather than guessed around):
      - Suggested H1 is always the same canned "Check manually as per ranking
        and traffic" note - that's onpage2.suggest_meta()'s own real behavior
        (it never generates a specific H1 replacement), not something
        simplified here.
      - "Missing Alt Text" gets a real per-image AI/heuristic ALT suggestion
        only when the page's "pages" entry has "missing_alt_srcs" (added to
        site_audit.py alongside this feature); a results dict produced by an
        older Site Audit run without that field falls back to a page-level
        note instead of guessing an image URL.
      - Sitemap/robots.txt issues are passed through as their own summary
        text, unchanged - same "never invent new claims about this" rule
        process_workbook() already follows for raw crawl-fact sheets.
      - No SE Ranking Audit sheet exists for JS/CSS issues, slow-loading
        pages, orphaned pages, crawl depth, canonical tags, or nofollow
        links - Site Audit doesn't currently check any of those, so there's
        nothing to map for them either way.
    """
    brand = (brand or results.get("domain") or "").strip()
    issues = results.get("issues") or {}
    page_lookup = _sa_page_lookup(results)
    cache = {}

    # Same pre-scan-then-set_run_scale() pattern process_workbook() uses, so
    # the AI paid-tier gate reflects this run's real scale before any
    # suggestion generation starts.
    total_rows = 0
    for _name, key, _builder in _SA_SHEET_SPECS:
        for it in issues.get(key) or []:
            total_rows += max(1, len(it.get("urls") or []))
    onpage2.set_run_scale(total_rows)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    issue_sheet_names = []

    for name, key, builder in _SA_SHEET_SPECS:
        items = issues.get(key) or []
        if not items:
            continue
        log(f"Generating '{name}' ({len(items)} row(s))...")
        headers, rows = builder(items, page_lookup, brand, cache)
        if not rows:
            continue
        _write_sheet(wb_out, name, headers, rows)
        issue_sheet_names.append(name)

    if issue_sheet_names:
        label = brand or results.get("domain") or Path(out_path).stem
        _write_all_issue_sheet(wb_out, issue_sheet_names, label)
    else:
        wb_out.create_sheet("No Issues Found").append(
            ["This site audit found no issues in a category this report covers."])

    wb_out.save(out_path)
    log(f"[DONE] {out_path}")
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_path", default=None, help="SEranking Site Audit .xlsx export")
    ap.add_argument("--pdf", dest="pdf_path", default=None, help="A PDF audit export (e.g. SEMrush overview)")
    ap.add_argument("--zip", dest="zip_path", default=None,
                     help="A zip of per-issue SEranking 'Pages' .xls exports")
    ap.add_argument("--out", dest="out_path", required=True, help="Output .xlsx path")
    ap.add_argument("--brand", default="", help="Brand name for suggested titles/descriptions")
    args = ap.parse_args()
    if not args.in_path and not args.pdf_path and not args.zip_path:
        ap.error("Provide --in (.xlsx) and/or --pdf and/or --zip")
    process_workbook(args.in_path, args.out_path, args.brand, args.pdf_path, args.zip_path)


if __name__ == "__main__":
    main()
